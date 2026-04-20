from __future__ import annotations
import hashlib
import io
import json
import math
import time
from dataclasses import dataclass, field
from typing import AsyncIterator

from app.infra.db.base import DBPool
from app.shared.exceptions import VocBaseError
from app.shared.logging import get_logger

logger = get_logger(__name__)

_CACHE_TTL = 300          # 5분
_CACHE_MAX_ENTRIES = 100
_CACHE_ROW_LIMIT = 10_000  # 이 이하면 캐싱, 초과면 offset 페이지네이션
_DOWNLOAD_MAX_ROWS = 100_000
_PAGE_SIZE_MAX = 200


class TableViewerError(VocBaseError):
    pass


@dataclass
class _CacheEntry:
    rows: list[dict]
    total: int
    expires_at: float


class TableService:
    def __init__(self, tc_pool: DBPool, whitelist: dict) -> None:
        self._pool = tc_pool
        self._tables: dict = whitelist.get("tables", {})
        self._cache: dict[str, _CacheEntry] = {}
        self._cache_order: list[str] = []  # LRU 순서

    # ── public ────────────────────────────────────────────────────────────────

    def list_tables(self) -> list[dict]:
        return [
            {
                "name": name,
                "columns": cfg.get("columns", []),
                "filterable": cfg.get("filterable", []),
                "requires_where": cfg.get("requires_where_clause", False),
            }
            for name, cfg in self._tables.items()
        ]

    async def get_rows(
        self,
        table: str,
        filters: dict[str, str],
        page: int,
        page_size: int,
    ) -> dict:
        cfg = self._get_config(table)
        page_size = min(page_size, _PAGE_SIZE_MAX)
        valid_filters = self._validate_filters(cfg, filters)

        if cfg.get("requires_where_clause") and not valid_filters:
            raise TableViewerError("이 테이블은 필터를 하나 이상 입력해야 합니다.")

        cache_key = self._make_key(table, valid_filters)
        cached = self._get_cache(cache_key)
        if cached:
            return self._slice(cached.rows, cached.total, page, page_size, from_cache=True)

        where_sql, params = self._build_where(valid_filters)
        cols_sql = self._cols_sql(cfg)

        count_rows = await self._pool.fetch_all(
            f"SELECT COUNT(*) AS cnt FROM `{table}` {where_sql}", params
        )
        total = int(count_rows[0]["cnt"])

        if total <= _CACHE_ROW_LIMIT:
            all_rows = await self._pool.fetch_all(
                f"SELECT {cols_sql} FROM `{table}` {where_sql} LIMIT {_CACHE_ROW_LIMIT}",
                params,
            )
            self._set_cache(cache_key, all_rows, total)
            return self._slice(all_rows, total, page, page_size, from_cache=False)

        offset = (page - 1) * page_size
        rows = await self._pool.fetch_all(
            f"SELECT {cols_sql} FROM `{table}` {where_sql} LIMIT {page_size} OFFSET {offset}",
            params,
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": math.ceil(total / page_size) if total else 1,
            "data": rows,
            "from_cache": False,
        }

    async def stream_download(
        self,
        table: str,
        filters: dict[str, str],
        fmt: str,
        limit: int | None,
    ) -> AsyncIterator[bytes]:
        cfg = self._get_config(table)
        valid_filters = self._validate_filters(cfg, filters)

        if cfg.get("requires_where_clause") and not valid_filters:
            raise TableViewerError("이 테이블은 필터를 하나 이상 입력해야 합니다.")

        columns = cfg.get("columns", [])
        where_sql, params = self._build_where(valid_filters)
        max_rows = min(limit, _DOWNLOAD_MAX_ROWS) if limit else _DOWNLOAD_MAX_ROWS
        cols_sql = self._cols_sql(cfg)

        if fmt == "csv":
            return self._stream_csv(table, columns, where_sql, params, cols_sql, max_rows)
        elif fmt == "excel":
            return self._stream_excel(table, columns, where_sql, params, cols_sql, max_rows)
        else:
            raise TableViewerError(f"지원하지 않는 포맷: {fmt}")

    # ── internals ─────────────────────────────────────────────────────────────

    def _get_config(self, table: str) -> dict:
        cfg = self._tables.get(table)
        if cfg is None:
            raise TableViewerError(f"테이블을 찾을 수 없습니다: {table}")
        return cfg

    def _validate_filters(self, cfg: dict, filters: dict[str, str]) -> dict[str, str]:
        allowed = set(cfg.get("filterable", []))
        return {k: v for k, v in filters.items() if k in allowed and v not in ("", None)}

    @staticmethod
    def _build_where(filters: dict[str, str]) -> tuple[str, dict]:
        if not filters:
            return "", {}
        parts = [f"`{col}` = %({col})s" for col in filters]
        return f"WHERE {' AND '.join(parts)}", dict(filters)

    @staticmethod
    def _cols_sql(cfg: dict) -> str:
        cols = cfg.get("columns", [])
        return ", ".join(f"`{c}`" for c in cols) if cols else "*"

    @staticmethod
    def _make_key(table: str, filters: dict) -> str:
        raw = json.dumps({"t": table, "f": sorted(filters.items())}, ensure_ascii=False)
        return hashlib.md5(raw.encode()).hexdigest()

    @staticmethod
    def _slice(rows: list, total: int, page: int, page_size: int, from_cache: bool) -> dict:
        start = (page - 1) * page_size
        end = start + page_size
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": math.ceil(total / page_size) if total else 1,
            "data": rows[start:end],
            "from_cache": from_cache,
        }

    # ── LRU cache ─────────────────────────────────────────────────────────────

    def _get_cache(self, key: str) -> _CacheEntry | None:
        entry = self._cache.get(key)
        if entry is None:
            return None
        if time.monotonic() > entry.expires_at:
            del self._cache[key]
            self._cache_order.remove(key)
            return None
        # LRU 갱신
        self._cache_order.remove(key)
        self._cache_order.append(key)
        return entry

    def _set_cache(self, key: str, rows: list, total: int) -> None:
        if key in self._cache:
            self._cache_order.remove(key)
        elif len(self._cache) >= _CACHE_MAX_ENTRIES:
            oldest = self._cache_order.pop(0)
            del self._cache[oldest]
        self._cache[key] = _CacheEntry(
            rows=rows, total=total, expires_at=time.monotonic() + _CACHE_TTL
        )
        self._cache_order.append(key)

    # ── streaming download ────────────────────────────────────────────────────

    async def _stream_csv(self, table, columns, where_sql, params, cols_sql, max_rows):
        import csv
        chunk_size = 1000
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(columns)
        yield buf.getvalue().encode("utf-8-sig")  # BOM for Excel 한글

        fetched = 0
        offset = 0
        while fetched < max_rows:
            batch = min(chunk_size, max_rows - fetched)
            rows = await self._pool.fetch_all(
                f"SELECT {cols_sql} FROM `{table}` {where_sql} LIMIT {batch} OFFSET {offset}",
                params,
            )
            if not rows:
                break
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row in rows:
                writer.writerow(list(row.values()))
            yield buf.getvalue().encode("utf-8-sig")
            fetched += len(rows)
            offset += len(rows)
            if len(rows) < batch:
                break

    async def _stream_excel(self, table, columns, where_sql, params, cols_sql, max_rows):
        from openpyxl import Workbook
        from openpyxl.styles import numbers as xl_numbers

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=table[:31])  # sheet name max 31 chars

        # 헤더
        ws.append(columns)

        fetched = 0
        offset = 0
        chunk_size = 1000
        while fetched < max_rows:
            batch = min(chunk_size, max_rows - fetched)
            rows = await self._pool.fetch_all(
                f"SELECT {cols_sql} FROM `{table}` {where_sql} LIMIT {batch} OFFSET {offset}",
                params,
            )
            if not rows:
                break
            for row in rows:
                cells = []
                for val in row.values():
                    cells.append(str(val) if val is not None else "")
                ws.append(cells)
            fetched += len(rows)
            offset += len(rows)
            if len(rows) < batch:
                break

        buf = io.BytesIO()
        wb.save(buf)

        # 텍스트 포맷 강제 (write_only 후 재오픈)
        buf.seek(0)
        from openpyxl import load_workbook
        wb2 = load_workbook(buf)
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.number_format = "@"
        buf2 = io.BytesIO()
        wb2.save(buf2)
        yield buf2.getvalue()
